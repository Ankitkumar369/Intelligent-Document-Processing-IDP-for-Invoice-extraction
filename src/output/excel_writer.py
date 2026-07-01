import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from src.core.config import OUTPUT_EXCEL_DIR, OUTPUT_EXCEL_FILENAME
from src.core.logger import log


class ExcelWriter:

    HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")

    def __init__(self):
        self.records = []
        self.item_records = []

    def add_record(self, normalized, doc_score, validation_report):
        try:
            self.records.append({
                "Filename": normalized.filename,
                "Invoice No": normalized.invoice_no,
                "Date": normalized.date_normalized,
                "Order ID": normalized.order_id,
                "Customer Name": normalized.customer_name,
                "Ship Mode": normalized.ship_mode,
                "Currency": normalized.currency,
                "Subtotal": normalized.subtotal,
                "Discount": normalized.discount,
                "Discount %": normalized.discount_pct,
                "Shipping": normalized.shipping,
                "Total": normalized.total,
                "Balance Due": normalized.balance_due,
                "Item Count": len(normalized.items),
                "Validation Passed": validation_report.passed,
                "Validation Score": validation_report.score,
                "Confidence Score": doc_score.overall_score,
                "Confidence Label": doc_score.overall_label,
            })

            for item in normalized.items:
                self.item_records.append({
                    "Invoice No": normalized.invoice_no,
                    "Filename": normalized.filename,
                    "Description": item.get("description", ""),
                    "Quantity": item.get("quantity", ""),
                    "Rate": item.get("rate", ""),
                    "Amount": item.get("amount", ""),
                    "Category": item.get("category", ""),
                    "Product Code": item.get("product_code", ""),
                })
        except Exception as e:
            log.error(f"ExcelWriter add_record error: {e}")

    def save(self):
        try:
            out_path = OUTPUT_EXCEL_DIR / OUTPUT_EXCEL_FILENAME

            df_summary = pd.DataFrame(self.records)
            df_items   = pd.DataFrame(self.item_records)

            with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
                df_summary.to_excel(writer, sheet_name="Invoices", index=False)
                if not df_items.empty:
                    df_items.to_excel(writer, sheet_name="Line Items", index=False)

                self._style_sheet(writer.sheets["Invoices"], df_summary)
                if not df_items.empty:
                    self._style_sheet(writer.sheets["Line Items"], df_items)

            log.info(f"OK Excel saved: {out_path} | rows={len(self.records)}")
            return str(out_path)

        except Exception as e:
            log.error(f"ExcelWriter save error: {e}")
            return ""

    def _style_sheet(self, ws, df):
        for col_num, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

            max_len = max(
                [len(str(col_name))] + [len(str(v)) for v in df[col_name].astype(str)]
            )
            ws.column_dimensions[get_column_letter(col_num)].width = min(max_len + 4, 40)

        ws.freeze_panes = "A2"
